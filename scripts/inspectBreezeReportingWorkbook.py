from pathlib import Path
from openpyxl import load_workbook

WORKBOOK_PATH = Path("/home/ubuntu/Downloads/breeze-intake/breeze-reporting.xlsx")
HEADER_HINTS = {
    "first name", "last name", "full name", "email", "email address", "phone",
    "city", "state", "date", "source", "campaign", "clicks", "visits",
}


def is_header(values: list[object]) -> bool:
    normalized = {str(value).strip().lower() for value in values if value not in (None, "")}
    return len(normalized.intersection(HEADER_HINTS)) >= 1


def main() -> None:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    print(f"workbook_tabs={len(workbook.sheetnames)}")
    for worksheet in workbook.worksheets:
        rows = [row for row in worksheet.iter_rows(values_only=True) if any(value not in (None, "") for value in row)]
        first_row = list(rows[0]) if rows else []
        header_present = is_header(first_row)
        data_rows = max(0, len(rows) - (1 if header_present else 0))
        max_columns = max((len(row) for row in rows), default=0)
        print(
            f"sheet={worksheet.title!r}; nonempty_rows={len(rows)}; "
            f"header_detected={header_present}; data_rows={data_rows}; columns={max_columns}"
        )
        if header_present:
            headers = [str(value).strip() for value in first_row if value not in (None, "")]
            print(f"headers={headers}")


if __name__ == "__main__":
    main()
