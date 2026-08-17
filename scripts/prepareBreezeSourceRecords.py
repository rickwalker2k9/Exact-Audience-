import csv
import json
from pathlib import Path

from openpyxl import load_workbook

REPORTING_WORKBOOK = Path("/home/ubuntu/Downloads/breeze-intake/breeze-reporting.xlsx")
EXACT_AUDIENCE_CSV = Path("/home/ubuntu/upload/LTDI_UNDER65_100KPLUS_PART_03.csv")
OUTPUT_PATH = Path("/home/ubuntu/Downloads/breeze-intake/breeze-source-records.ndjson")
GOOGLE_ALLOCATION = 112


def clean(value):
    return "" if value is None else str(value).strip()


def export_reporting_records(handle):
    workbook = load_workbook(REPORTING_WORKBOOK, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = [list(row) for row in worksheet.iter_rows(values_only=True) if any(value not in (None, "") for value in row)]
    headers = [clean(value).upper() for value in rows[0]]
    column = {name: index for index, name in enumerate(headers)}

    def cell(row, name):
        index = column.get(name)
        return clean(row[index]) if index is not None and index < len(row) else ""

    count = 0
    for index, row in enumerate(rows[1:], start=1):
        source = "google-ads" if index <= GOOGLE_ALLOCATION else "meta-ads"
        record = {
            "source": source,
            "firstName": cell(row, "FIRST NAME"),
            "lastName": cell(row, "LAST NAME"),
            "email": cell(row, "VERFIED EMAIL") or cell(row, "VERIFIED EMAIL"),
            "phone": cell(row, "PHONE"),
            "ageRange": "",
            "incomeRange": "",
            "city": cell(row, "CITY"),
            "state": cell(row, "ST"),
            "sourceLabel": "Google Ads · user-provided allocation" if source == "google-ads" else "Meta Ads · user-provided allocation",
            "recordOrdinal": index,
        }
        handle.write(json.dumps(record) + "\n")
        count += 1
    return count


def export_exact_audience_records(handle):
    count = 0
    with EXACT_AUDIENCE_CSV.open("r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.DictReader(csv_file)
        for index, row in enumerate(reader, start=1):
            record = {
                "source": "exact-audience",
                "firstName": clean(row.get("First Name")),
                "lastName": clean(row.get("Last Name")),
                "email": clean(row.get("Personal Verified Emails")),
                "phone": clean(row.get("Skiptrace Wireless Numbers")),
                "ageRange": clean(row.get("Age Range")),
                "incomeRange": clean(row.get("Income Range")),
                "city": clean(row.get("Personal City")),
                "state": clean(row.get("Personal State")),
                "sourceLabel": "Exact Audience · Email Outreach",
                "recordOrdinal": index,
            }
            handle.write(json.dumps(record) + "\n")
            count += 1
    return count


def main():
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_PATH.open("w", encoding="utf-8") as handle:
        reporting_count = export_reporting_records(handle)
        exact_audience_count = export_exact_audience_records(handle)
    print(f"reporting_records={reporting_count}")
    print(f"google_ads_records={GOOGLE_ALLOCATION}")
    print(f"meta_ads_records={reporting_count - GOOGLE_ALLOCATION}")
    print(f"exact_audience_records={exact_audience_count}")
    print(f"output={OUTPUT_PATH}")


if __name__ == "__main__":
    main()
